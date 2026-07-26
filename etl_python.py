"""
ETL Pipeline: Dampak IKN terhadap Ekonomi & Migrasi Penduduk Kaltim
====================================================================
Extract  -> baca file mentah dari data/raw/
Transform -> cleaning, unpivot wide->long, fix known data issues
Load     -> simpan dataset bersih (star schema) ke data/processed/

Jalankan: python3 etl_pipeline.py
"""

import pandas as pd
import openpyxl
from pathlib import Path

RAW_DIR = Path("data/raw")
OUT_DIR = Path("data/processed")
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Daftar 10 kabupaten/kota yang valid (dipakai untuk memisahkan baris
# agregat provinsi "Kalimantan Timur" dari baris kabupaten/kota asli)
KABKOTA_VALID = [
    "Paser", "Kutai Barat", "Kutai Kartanegara", "Kutai Timur", "Berau",
    "Penajam Paser Utara", "Mahakam Ulu", "Balikpapan", "Samarinda", "Bontang",
]

# Kategorisasi wilayah berdasarkan keterkaitan fungsional dengan pembangunan IKN
# (bukan cuma kedekatan geografis, tapi peran ekonomi/logistik)
KATEGORI_WILAYAH = {
    "Penajam Paser Utara": "IKN Inti",
    "Kutai Kartanegara": "IKN Inti",
    "Paser": "Penyangga",
    "Balikpapan": "Penyangga",       # akses logistik utama (bandara, pelabuhan)
    "Samarinda": "Penyangga",        # pusat pemerintahan & ekonomi provinsi
    "Kutai Barat": "Non-IKN",
    "Kutai Timur": "Non-IKN",
    "Berau": "Non-IKN",
    "Bontang": "Non-IKN",
    "Mahakam Ulu": "Non-IKN",
}


def _clean_kabkota_name(name: str) -> str:
    """Samakan penamaan, mis. 'Kota Balikpapan' -> 'Balikpapan'."""
    return name.replace("Kota ", "").replace("Kabupaten ", "").strip()


def _clean_numeric_value(value):
    """BPS kadang menulis '-' untuk data yang tidak tersedia/tidak dipublikasikan
    di level kab/kota (mis. TPT 2016) -> jadi None. Beberapa file juga menyimpan
    angka sebagai teks (mis. '160.9' atau '160,9') alih-alih angka asli -> convert
    ke float. Ini penting supaya perbandingan numerik (mis. di fix_known_issues)
    tidak error karena tipe data bercampur str dan angka."""
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        if cleaned in ("-", "", "..", "..."):
            return None
        cleaned = cleaned.replace(",", ".")  # format Indonesia: koma = desimal
        try:
            return float(cleaned)
        except ValueError:
            return None
    return value


# ---------------------------------------------------------------------------
# 1. EXTRACT + TRANSFORM: PDRB per Kapita (ADHB & ADHK)
# ---------------------------------------------------------------------------
def load_pdrb() -> pd.DataFrame:
    all_frames = []
    for path in sorted((RAW_DIR / "pdrb").glob("*.xlsx")):
        wb = openpyxl.load_workbook(path, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            # Deteksi ADHB/ADHK dari teks judul (bukan asumsi urutan sheet),
            # karena urutan sheet bisa berbeda antar file
            judul = str(rows[0][1]) if len(rows[0]) > 1 and rows[0][1] else ""
            if "Harga Berlaku" in judul:
                harga_tipe = "adhb"
            elif "Harga Konstan" in judul:
                harga_tipe = "adhk"
            else:
                print(f"[WARNING] Sheet '{sheet_name}' di {path.name} tidak dikenali (ADHB/ADHK), dilewati. Judul: {judul!r}")
                continue

            years = [y for y in rows[1] if isinstance(y, int)]

            for row in rows[2:]:
                kabkota = row[0]
                if kabkota is None or kabkota not in KABKOTA_VALID:
                    continue
                for i, year in enumerate(years):
                    all_frames.append({
                        "kabupaten_kota": _clean_kabkota_name(kabkota),
                        "tahun": year,
                        "harga_tipe": harga_tipe,
                        "pdrb_per_kapita_juta_rp": _clean_numeric_value(row[i + 1]),
                    })

    df = pd.DataFrame(all_frames).drop_duplicates(subset=["kabupaten_kota", "tahun", "harga_tipe"])
    # pivot supaya adhb & adhk jadi 2 kolom, bukan 2 baris terpisah
    df = df.pivot_table(
        index=["kabupaten_kota", "tahun"],
        columns="harga_tipe",
        values="pdrb_per_kapita_juta_rp",
    ).reset_index()
    df.columns.name = None
    df = df.rename(columns={
        "adhb": "pdrb_adhb_juta_rp",
        "adhk": "pdrb_adhk_juta_rp",
    })
    return df


# ---------------------------------------------------------------------------
# 2. EXTRACT + TRANSFORM: Tenaga Kerja (TPT & TPAK)
# ---------------------------------------------------------------------------
def load_tenaga_kerja() -> pd.DataFrame:
    all_frames = []
    for path in sorted((RAW_DIR / "tenaga_kerja").glob("*.xlsx")):
        wb = openpyxl.load_workbook(path, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            # Deteksi indikator dari TEKS JUDUL (bukan asumsi urutan sheet),
            # karena urutan 'Data 1'/'Data 2' ternyata bisa terbalik antar file
            # (mis. file 2015-2018 punya urutan TPAK dulu baru TPT, kebalik
            # dari file 2019-2024). Deteksi berbasis teks judul lebih aman.
            judul = str(rows[0][1]) if len(rows[0]) > 1 and rows[0][1] else ""
            if "Pengangguran" in judul:
                indikator = "tpt_persen"
            elif "Partisipasi" in judul:
                indikator = "tpak_persen"
            else:
                print(f"[WARNING] Sheet '{sheet_name}' di {path.name} tidak dikenali indikatornya, dilewati. Judul: {judul!r}")
                continue

            years = [y for y in rows[1] if isinstance(y, int)]

            for row in rows[2:]:
                kabkota = row[0]
                if kabkota is None or kabkota not in KABKOTA_VALID:
                    continue
                for i, year in enumerate(years):
                    all_frames.append({
                        "kabupaten_kota": _clean_kabkota_name(kabkota),
                        "tahun": year,
                        "indikator": indikator,
                        "nilai": _clean_numeric_value(row[i + 1]),
                    })

    df = pd.DataFrame(all_frames).drop_duplicates(subset=["kabupaten_kota", "tahun", "indikator"])
    df = df.pivot_table(
        index=["kabupaten_kota", "tahun"], columns="indikator", values="nilai"
    ).reset_index()
    df.columns.name = None
    return df


# ---------------------------------------------------------------------------
# 3. EXTRACT + TRANSFORM: Kependudukan (6 file terpisah per tahun)
# ---------------------------------------------------------------------------
def load_kependudukan() -> pd.DataFrame:
    frames = []
    for path in sorted((RAW_DIR / "kependudukan").glob("penduduk_*.xlsx")):
        tahun = int(path.stem.split("_")[1])
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["Sheet1"]
        rows = list(ws.iter_rows(values_only=True))

        for row in rows[1:]:
            kabkota_raw = row[0]
            if kabkota_raw is None:
                continue
            kabkota = _clean_kabkota_name(kabkota_raw)  # "Kota Balikpapan" -> "Balikpapan"
            if kabkota not in KABKOTA_VALID:
                continue  # skip "Kalimantan Timur" (agregat) & baris metadata
            frames.append({
                "kabupaten_kota": kabkota,
                "tahun": tahun,
                "jumlah_penduduk_ribu": _clean_numeric_value(row[1]),
                "kepadatan_penduduk_km2": _clean_numeric_value(row[4]),
                "rasio_jenis_kelamin": _clean_numeric_value(row[5]),
                # kolom laju pertumbuhan bawaan BPS TIDAK dipakai di sini karena
                # basis perhitungannya berubah-ubah antar tahun (lihat README) -
                # kita hitung ulang sendiri di fungsi fix_known_issues()
            })
    return pd.DataFrame(frames)


# ---------------------------------------------------------------------------
# 3b. EXTRACT + TRANSFORM: Investasi (PMDN & PMA) - opsional, cakupan tahun
#     terbatas (2015-2023, belum ada 2024 saat data ini diambil). PMDN dan PMA
#     TIDAK digabung karena satuan mata uang berbeda (PMDN = Juta Rupiah,
#     PMA = Ribu US Dollar) - digabung tanpa konversi kurs akan salah secara
#     matematis. Keduanya dianalisis terpisah sebagai indikator pelengkap.
#
#     PENTING: berbeda dengan penduduk & PDRB, growth rate (%) TIDAK dihitung
#     untuk investasi. Data realisasi investasi bersifat 'lumpy' (bisa
#     melonjak >40x dalam 1 tahun karena 1 proyek besar disetujui, atau
#     bernilai 0 di tahun tanpa realisasi) - ini membuat persentase growth
#     rate menghasilkan angka ekstrem/tidak terhingga yang tidak representatif,
#     bahkan setelah dirata-ratakan pakai median sekalipun. Sebagai gantinya,
#     investasi dianalisis dengan membandingkan RATA-RATA NILAI NOMINAL per
#     tahun (before vs after) - pola yang sama seperti TPT/TPAK.
# ---------------------------------------------------------------------------
def load_investasi() -> pd.DataFrame:
    investasi_dir = RAW_DIR / "investasi"
    if not investasi_dir.exists():
        return pd.DataFrame(columns=["kabupaten_kota", "tahun", "pmdn_juta_rp", "pma_ribu_usd"])

    all_frames = []
    for path in sorted(investasi_dir.glob("*.xlsx")):
        wb = openpyxl.load_workbook(path, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            judul = str(rows[0][1]) if len(rows[0]) > 1 and rows[0][1] else ""
            if "PMDN" in judul:
                indikator = "pmdn_juta_rp"
            elif "PMA" in judul:
                indikator = "pma_ribu_usd"
            else:
                print(f"[WARNING] Sheet '{sheet_name}' di {path.name} tidak dikenali (PMDN/PMA), dilewati. Judul: {judul!r}")
                continue

            years = [y for y in rows[1] if isinstance(y, int)]

            for row in rows[2:]:
                kabkota = row[0]
                if kabkota is None or kabkota not in KABKOTA_VALID:
                    continue
                for i, year in enumerate(years):
                    all_frames.append({
                        "kabupaten_kota": _clean_kabkota_name(kabkota),
                        "tahun": year,
                        "indikator": indikator,
                        "nilai": _clean_numeric_value(row[i + 1]),
                    })

    if not all_frames:
        return pd.DataFrame(columns=["kabupaten_kota", "tahun", "pmdn_juta_rp", "pma_ribu_usd"])

    df = pd.DataFrame(all_frames).drop_duplicates(subset=["kabupaten_kota", "tahun", "indikator"])
    df = df.pivot_table(
        index=["kabupaten_kota", "tahun"], columns="indikator", values="nilai"
    ).reset_index()
    df.columns.name = None
    return df


def fix_known_issues(df_penduduk: pd.DataFrame) -> pd.DataFrame:
    """
    Masalah #1: Samarinda 2021 'jumlah_penduduk_ribu' = 83.1 (typo sumber BPS,
    seharusnya ~831). Kita perbaiki dengan interpolasi linear dari 2020 & 2022.
    """
    mask = (df_penduduk["kabupaten_kota"] == "Samarinda") & (df_penduduk["tahun"] == 2021)
    if mask.any() and df_penduduk.loc[mask, "jumlah_penduduk_ribu"].iloc[0] < 100:
        pop_2020 = df_penduduk.query("kabupaten_kota == 'Samarinda' and tahun == 2020")["jumlah_penduduk_ribu"].iloc[0]
        pop_2022 = df_penduduk.query("kabupaten_kota == 'Samarinda' and tahun == 2022")["jumlah_penduduk_ribu"].iloc[0]
        interpolated = round((pop_2020 + pop_2022) / 2, 1)
        df_penduduk.loc[mask, "jumlah_penduduk_ribu"] = interpolated
        print(f"[FIX] Samarinda 2021: 83.1 -> {interpolated} (interpolasi 2020 & 2022)")

    return df_penduduk


def compute_growth_rate(df_penduduk: pd.DataFrame) -> pd.DataFrame:
    """Hitung ulang laju pertumbuhan penduduk (%) secara konsisten,
    menggantikan kolom bawaan BPS yang basis hitungnya tidak seragam."""
    df_penduduk = df_penduduk.sort_values(["kabupaten_kota", "tahun"])
    df_penduduk["laju_pertumbuhan_persen"] = (
        df_penduduk.groupby("kabupaten_kota")["jumlah_penduduk_ribu"].pct_change() * 100
    ).round(2)
    return df_penduduk


def compute_pdrb_growth_rate(fact: pd.DataFrame) -> pd.DataFrame:
    """Hitung growth rate PDRB (ADHK - harga konstan, supaya efek inflasi
    sudah dihilangkan) per kab/kota per tahun, dengan pola yang sama seperti
    laju_pertumbuhan_persen untuk penduduk. Dipakai untuk perbandingan
    before-after dampak IKN terhadap pertumbuhan ekonomi."""
    fact = fact.sort_values(["kabupaten_kota", "tahun"])
    fact["pdrb_growth_persen"] = (
        fact.groupby("kabupaten_kota")["pdrb_adhk_juta_rp"].pct_change() * 100
    ).round(2)
    return fact


# ---------------------------------------------------------------------------
# 5. MAIN: gabungkan semua jadi satu fact table + dimensi
# ---------------------------------------------------------------------------
def main():
    print("Extracting & transforming data...")
    df_pdrb = load_pdrb()
    df_tenaga_kerja = load_tenaga_kerja()
    df_penduduk = load_kependudukan()
    df_penduduk = fix_known_issues(df_penduduk)
    df_penduduk = compute_growth_rate(df_penduduk)
    df_investasi = load_investasi()

    # --- Fact table: gabungan semua indikator per kab/kota per tahun ---
    fact = (
        df_pdrb
        .merge(df_tenaga_kerja, on=["kabupaten_kota", "tahun"], how="outer")
        .merge(df_penduduk, on=["kabupaten_kota", "tahun"], how="outer")
        .merge(df_investasi, on=["kabupaten_kota", "tahun"], how="left")  # left join: investasi cakupan tahun lebih pendek (s.d. 2023)
    )
    fact = compute_pdrb_growth_rate(fact)
    fact["kategori_wilayah"] = fact["kabupaten_kota"].map(KATEGORI_WILAYAH)
    fact["periode_ikn"] = fact["tahun"].apply(lambda t: "sebelum" if t < 2019 else ("pengumuman" if t == 2019 else "sesudah"))

    # --- Dimension table: kabupaten/kota ---
    dim_kabkota = pd.DataFrame({
        "kabupaten_kota": KABKOTA_VALID,
        "kategori_wilayah": [KATEGORI_WILAYAH[k] for k in KABKOTA_VALID],
    })

    # --- Save ke CSV (bisa langsung di-load ke Postgres / Power BI) ---
    fact_path = OUT_DIR / "fact_regional_indicator.csv"
    dim_path = OUT_DIR / "dim_kabupaten_kota.csv"
    fact.sort_values(["tahun", "kabupaten_kota"]).to_csv(fact_path, index=False)
    dim_kabkota.to_csv(dim_path, index=False)

    print(f"\nSelesai. {len(fact)} baris fact table disimpan ke:")
    print(f"  - {fact_path}")
    print(f"  - {dim_path}")
    print("\nPreview fact table:")
    print(fact.head(10).to_string(index=False))

    # --- Quick data quality check ---
    print("\n=== Cek missing values per kolom ===")
    print(fact.isna().sum())


if __name__ == "__main__":
    main()